import subprocess
import openpyxl as oxl
'''
エクセルは閉じて実行すること
途中に空欄があるとうまく動作しない
塗り始めの行列は左上を指定する
色を変えたい、色の範囲を変えたい場合はしたの方にある関数呼び出しのところを変える
'''
# 色を塗りたいエクセルのファイルとシートを参照
path = "./csv/insibunseki.xlsx"
book = oxl.load_workbook(path)
sheet = book["Sheet1"]
#エクセルアプリのパス
excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
#塗り始める行と列指定
start_row = 5
start_column = ord("W")-64
# 引数1=塗り始める行  引数2=塗り始める列  引数3=塗りたい色RGB255で指定  引数4=塗りたい値の最大値  引数5=塗りたい値の最小値
def paint_cell_bg(start_row, start_column, rgb_255, corr_max, corr_min):
    color_fill = oxl.styles.PatternFill(patternType="solid",fgColor=rgb_255,bgColor=rgb_255)
    print(f"{corr_max}以下,{corr_min}以上の値をペイントします")
    pre_value = None
    i = start_row
    j = start_column
    while True:
        if sheet.cell(row=i,column=j).value == None:
            j += 1
            i = start_row - 1
            if pre_value == None:
                break
        elif corr_max >= sheet.cell(row=i,column=j).value > corr_min or (-1)*corr_min >= sheet.cell(row=i,column=j).value > (-1)*corr_max:
            sheet.cell(row=i,column=j).fill = color_fill
        pre_value = sheet.cell(row=i,column=j).value
        i += 1
#黄色
rgb_255 = "ffff33"
paint_cell_bg(start_row,start_column,rgb_255,0.4,0.1)
#オレンジ
rgb_255 = "ffcc00"
paint_cell_bg(start_row,start_column,rgb_255,0.7,0.4)
#赤
rgb_255 = "ff0000"
paint_cell_bg(start_row,start_column,rgb_255,100,0.7)

print("ペイント終了")
book.save(path)
subprocess.Popen([excel_path,path])
